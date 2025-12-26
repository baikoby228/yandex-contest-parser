from openpyxl import load_workbook

from person import Person
from excel_writer import write_to_excel

def merge(file_names: list, task_count: int, row=None) -> list:
    names = set()
    for file_name in file_names:
        wb = load_workbook(file_name + '.xlsx')
        sheet = wb.active
        for row in sheet:
            name = row[1].value
            names.add(name)

    cnt = {}
    for file_name in file_names:
        wb = load_workbook(file_name + '.xlsx')
        sheet = wb.active

        used = []
        f = True
        for row in sheet:
            if f:
                f = False
                continue

            name = row[1].value
            total = row[task_count + 2].value

            if name in used:
                continue
            used.append(name)

            if name not in cnt:
                cnt[name] = []
            cnt[name].append(total)

        for name in names:
            if name not in used:
                if name not in cnt:
                    cnt[name] = []
                cnt[name].append(0)

    res = []
    cnt_points = {}
    for key, value in cnt.items():
        person = Person()
        person.name = key

        for x in value:
            person.tasks_count += 1
            person.points.append(x)
            person.total_points += x

        if person.total_points not in cnt_points:
            cnt_points[person.total_points] = 0
        cnt_points[person.total_points] += 1

        res.append(person)
    res.sort(key=lambda x: x.total_points, reverse=True)

    last = -1
    cnt_points[last] = 0

    cur = 1
    for person in res:
        if last != person.total_points:
            cur += cnt_points[last]
        if cnt_points[person.total_points] == 1:
            person.place = str(cur)
        else:
            person.place = f'{cur}-{cur + cnt_points[person.total_points] - 1}'
        last = person.total_points

    return res

if __name__ == '__main__':
    persons = merge(['informatika2', 'informatika3'], 8)
    write_to_excel(persons, '2-3', False, 2)
